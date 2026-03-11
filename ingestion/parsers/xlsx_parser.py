"""Parser for .xlsx files using openpyxl.

Treats the first row of each sheet as a header.
Each subsequent row becomes a text block in the format:
  "Header1: val | Header2: val | ..."

Each returned dict has the shape:
  {
      "text": str,
      "source": str,        # filename
      "section": str,       # sheet name
      "page": int,          # 1-based data row number (excluding header)
  }

TODO: Handle merged cells — openpyxl exposes merged cell ranges via
      `sheet.merged_cells.ranges`. Currently merged cells may produce
      empty or repeated values. Consider un-merging and propagating
      the top-left cell value before iterating rows.
"""

import logging
from pathlib import Path

import openpyxl

logger = logging.getLogger(__name__)


def parse(file_path: Path) -> list[dict[str, str | int | None]]:
    """Extract text chunks from an .xlsx file.

    Args:
        file_path: Absolute path to the .xlsx file.

    Returns:
        List of chunk dicts ready for the chunker.
    """
    workbook = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    source = file_path.name
    chunks: list[dict[str, str | int | None]] = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            logger.debug("xlsx_parser: sheet '%s' in '%s' is empty, skipping", sheet_name, source)
            continue

        headers = [str(h).strip() if h is not None else f"Col{i}" for i, h in enumerate(rows[0])]

        for row_idx, row in enumerate(rows[1:], start=1):
            parts = []
            for header, value in zip(headers, row):
                if value is not None and str(value).strip():
                    parts.append(f"{header}: {str(value).strip()}")
            row_text = " | ".join(parts)
            if row_text:
                chunks.append(
                    {
                        "text": row_text,
                        "source": source,
                        "section": sheet_name,
                        "page": row_idx,
                    }
                )

    logger.info("xlsx_parser: extracted %d chunks from '%s'", len(chunks), source)
    return chunks
